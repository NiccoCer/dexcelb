# ===== COPIA TUTTO QUESTO FILE =====

import json
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# ===== COSTANTI (mantieni uguali al tuo programma originale) =====
COL_CONVERTITA = "CONVERTITA"
COL_NON_CONV = "PASSATA NON CONVERTITA"
MERGE_KEY_COLS = ["NOME", "COGNOME", "ZONA"]

DEFAULT_TEMPLATE_CONVERTITA = (
    "-{NOME} {COGNOME};\n"
    "-{TELEFONO};\n"
    "-{MQ};\n"
    "-{INDIRIZZO};\n"
    "(Già chiamato, si aspetta una chiamata in giornata)"
)

DEFAULT_TEMPLATE_NON_CONV = (
    "-{NOME} {COGNOME};\n"
    "-{TELEFONO};\n"
    "-{MQ};\n"
    "-{INDIRIZZO};\n"
    "Passata non convertita, continuiamo a provare a contattarla."
)


# ===== SERIALIZZAZIONE WORKBOOK =====
def serialize_workbook(wb):
    """Salva workbook in JSON per la sessione Flask"""
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "headers": [cell.value for cell in ws[1]],
            "rows": []
        }
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:  # Salta header
                continue
            sheet_data["rows"].append(
                [str(c) if c is not None else "" for c in row]
            )
        data[sheet_name] = sheet_data
    return json.dumps(data)


def deserialize_workbook(serialized_data):
    """Ricostruisce workbook da JSON"""
    data = json.loads(serialized_data)
    wb = Workbook()
    ws = wb.active

    sheet_name = next(iter(data.keys()))
    sheet = data[sheet_name]

    headers = sheet["headers"] or []
    rows = sheet["rows"] or []

    # Scrivi headers
    for col, header in enumerate(headers, start=1):
        ws[f"{get_column_letter(col)}1"] = header

    # Scrivi righe
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, cell_data in enumerate(row_data, start=1):
            ws[f"{get_column_letter(col_idx)}{row_idx}"] = cell_data

    return wb


# ===== FUNZIONI UTILITÀ DB =====
def get_headers(wb):
    """Ritorna lista headers dal primo foglio"""
    ws = wb.active
    return [str(cell.value) if cell.value else "" for cell in ws[1]]


def find_column_index(ws, header_name):
    """Trova indice colonna (1-based) per nome header"""
    if ws.max_column is None:
        return None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if str(val).strip().upper() == str(header_name).strip().upper():
            return col
    return None


def mark_row_status(wb, row_idx, status):
    """Aggiorna stato riga: 'convertita', 'non_conv', 'clear'"""
    ws = wb.active
    col_conv = find_column_index(ws, COL_CONVERTITA)
    col_non_conv = find_column_index(ws, COL_NON_CONV)

    if not col_conv or not col_non_conv:
        return wb

    if status == "clear":
        ws.cell(row=row_idx, column=col_conv).value = ""
        ws.cell(row=row_idx, column=col_non_conv).value = ""
    elif status == "convertita":
        ws.cell(row=row_idx, column=col_conv).value = "X"
        ws.cell(row=row_idx, column=col_non_conv).value = ""
    elif status == "non_conv":
        ws.cell(row=row_idx, column=col_conv).value = ""
        ws.cell(row=row_idx, column=col_non_conv).value = "X"

    return wb


def generate_copy_text(wb, row_idx, templates):
    """Genera testo formattato per riga (copia negli appunti)"""
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    mappa = {}
    for i, h in enumerate(headers):
        key = (str(h) if h else "").strip().upper()
        if not key:
            continue
        mappa[key] = str(ws.cell(row=row_idx, column=i + 1).value or "").strip()

    conv_val = mappa.get(COL_CONVERTITA.upper(), "").strip()
    non_conv_val = mappa.get(COL_NON_CONV.upper(), "").strip()

    if conv_val:
        template = templates["convertita"]
    elif non_conv_val:
        template = templates["non_conv"]
    else:
        template = templates["non_conv"]

    return template.format(
        NOME=mappa.get("NOME", ""),
        COGNOME=mappa.get("COGNOME", ""),
        TELEFONO=mappa.get("TELEFONO", ""),
        MQ=mappa.get("MQ", ""),
        INDIRIZZO=mappa.get("INDIRIZZO", ""),
    )


def add_row_to_workbook(wb, row_values):
    """Aggiunge una riga al workbook"""
    ws = wb.active
    ws.append(row_values)
    return wb


def delete_row_from_workbook(wb, row_idx):
    """Elimina una riga dal workbook (row_idx è 1-based di Excel)"""
    ws = wb.active
    if row_idx > 1 and row_idx <= ws.max_row:
        ws.delete_rows(row_idx, 1)
    return wb


def merge_workbooks(wb_master, wb_source, merge_cols):
    """
    Unisce wb_source in wb_master.
    merge_cols: lista colonne su cui deduplicare (es. ['NOME', 'COGNOME', 'ZONA']).
    """
    ws_master = wb_master.active
    ws_source = wb_source.active

    # Headers master
    headers_master = [cell.value for cell in ws_master[1]]

    # Raccogli chiavi deduplica da master
    master_keys = set()
    for row_idx in range(2, ws_master.max_row + 1):
        key_vals = []
        for col_name in merge_cols:
            col_idx = find_column_index(ws_master, col_name)
            if col_idx:
                val = ws_master.cell(row=row_idx, column=col_idx).value or ""
                key_vals.append(str(val).strip().upper())
        if key_vals:
            master_keys.add(tuple(key_vals))

    # Aggiungi righe da source se non duplicate
    righe_importate = 0
    for row_idx in range(2, ws_source.max_row + 1):
        key_vals = []
        for col_name in merge_cols:
            col_idx = find_column_index(ws_source, col_name)
            if col_idx:
                val = ws_source.cell(row=row_idx, column=col_idx).value or ""
                key_vals.append(str(val).strip().upper())

        if tuple(key_vals) not in master_keys:
            row_data = []
            for col_idx in range(1, ws_source.max_column + 1):
                row_data.append(ws_source.cell(row=row_idx, column=col_idx).value)
            ws_master.append(row_data)
            master_keys.add(tuple(key_vals))
            righe_importate += 1

    return wb_master, righe_importate


# ===== FINE api/utils.py =====

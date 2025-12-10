from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
import os
import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
import json

# --- COSTANTI ---

COL_CONVERTITA = "CONVERTITA"
COL_NON_CONV = "PASSATA NON CONVERTITA"
MERGE_KEY_COLS = ["NOME", "COGNOME", "ZONA"]
DB_PATH = "db_master.xlsx" # Percorso fisso per il DB su Vercel
CONFIG_FILE = "dexcelb_templates.json"

DEFAULT_TEMPLATE_CONVERTITA = (
    "-{NOME} {COGNOME};\n"
    "-{TELEFONO};\n"
    "-{MQ};\n"
    "-{INDIRIZZO};\n"
    "(Già chiamato, si aspetta una chiamata in giornata)"
)

DEFAULT_TEMPLATE_NON_CONV = (
    "-{NOME} {COGNOME};\n"
    "-{TELEFONO};\n"
    "-{MQ};\n"
    "-{INDIRIZZO};\n"
    "Passata non convertita, continuiamo a provare a contattarla."
)

# --- FUNZIONI DI UTILITÀ EXCEL ---

def trova_indice_colonna(ws, header_name):
    num_col = ws.max_column or 0
    for i in range(1, num_col + 1):
        val = ws.cell(row=1, column=i).value
        if str(val).strip().lower() == str(header_name).strip().lower():
            return i
    return None

def carica_dati_da_file(filepath):
    if not os.path.exists(filepath):
        return [], []

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    colonne_attuali = [
        str(h) if h not in (None, "") else f"Colonna {i+1}"
        for i, h in enumerate(headers)
    ]
    num_col = len(colonne_attuali)

    righe = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row = list(row) if row is not None else []
        if len(row) > num_col:
            row = row[:num_col]
        elif len(row) < num_col:
            row += [None] * (num_col - len(row))
        
        righe.append({'riga_excel': row_idx, 'valori': [str(v) if v is not None else '' for v in row]})
        
    return colonne_attuali, righe

def importa_dati_da_buffer(db_path, file_content):
    if not os.path.exists(db_path):
         raise FileNotFoundError("DB master non trovato.")
    
    try:
        wb_db = load_workbook(db_path)
        ws_db = wb_db.active

        wb_src = load_workbook(io.BytesIO(file_content), data_only=True)
        ws_src = wb_src.active
    except InvalidFileException:
        raise ValueError("File Excel sorgente non valido.")
    
    righe_copiate = 0
    for i, row in enumerate(ws_src.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if all(cell is None for cell in row):
            continue
        # Assicurati che le righe abbiano lo stesso numero di colonne del master
        master_cols = wb_db.active.max_column
        row_list = list(row)
        if len(row_list) > master_cols:
            row_list = row_list[:master_cols]
        elif len(row_list) < master_cols:
            row_list += [None] * (master_cols - len(row_list))
            
        ws_db.append(row_list)
        righe_copiate += 1

    wb_db.save(db_path)
    return righe_copiate

def unisci_file_lista(master_path, file_content_list):
    if not os.path.exists(master_path):
        raise FileNotFoundError(f"File master non trovato: {master_path}")

    wb_master = load_workbook(master_path)
    ws_master = wb_master.active

    master_headers = [cell.value for cell in ws_master[1]]
    num_col_master = len(master_headers)

    header_to_index = {
        str(h).strip().upper(): i
        for i, h in enumerate(master_headers)
        if h is not None
    }
    
    def costruisci_chiave(row):
        valori = []
        for col in MERGE_KEY_COLS:
            idx = header_to_index.get(col.upper(), None)
            if idx is None or idx >= len(row):
                valori.append("")
            else:
                v = row[idx]
                valori.append(str(v or "").strip().upper())
        return "|".join(valori)

    chiavi = set()
    for i, row in enumerate(ws_master.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        row = list(row) if row is not None else []
        if all(cell is None for cell in row):
            continue
        key = costruisci_chiave(row)
        if key.strip():
            chiavi.add(key)

    nuovi = 0
    files_trovati = 0

    for file_content in file_content_list:
        files_trovati += 1
        
        try:
            wb_src = load_workbook(io.BytesIO(file_content), data_only=True)
            ws_src = wb_src.active
        except InvalidFileException:
            continue

        for i, row in enumerate(ws_src.iter_rows(values_only=True), start=1):
            if i == 1:
                continue
            row = list(row) if row is not None else []
            if all(cell is None for cell in row):
                continue

            if len(row) > num_col_master:
                row = row[:num_col_master]
            elif len(row) < num_col_master:
                row += [None] * (num_col_master - len(row))

            key = costruisci_chiave(row)

            if not key.strip():
                continue
            if key in chiavi:
                continue

            ws_master.append(row)
            chiavi.add(key)
            nuovi += 1

    wb_master.save(master_path)
    return files_trovati, nuovi

def aggiorna_stato_riga(db_path, riga_excel, convertita=False, non_convertita=False, pulisci=False):
    if riga_excel == 1:
        raise ValueError("Non è possibile modificare la riga di intestazione.")
    
    if not os.path.exists(db_path):
        raise FileNotFoundError("DB master non trovato.")
        
    wb = load_workbook(db_path)
    ws = wb.active

    col_conv = trova_indice_colonna(ws, COL_CONVERTITA)
    col_non_conv = trova_indice_colonna(ws, COL_NON_CONV)

    if col_conv is None or col_non_conv is None:
        raise KeyError(f"Colonne di stato mancanti: '{COL_CONVERTITA}' e '{COL_NON_CONV}' sono necessarie nella prima riga.")

    if pulisci:
        ws.cell(row=riga_excel, column=col_conv).value = ""
        ws.cell(row=riga_excel, column=col_non_conv).value = ""
    else:
        ws.cell(row=riga_excel, column=col_conv).value = "X" if convertita else ""
        ws.cell(row=riga_excel, column=col_non_conv).value = "X" if non_convertita else ""

    wb.save(db_path)
    return True

def aggiungi_riga_manuale(db_path, row_values):
    if not os.path.exists(db_path):
        raise FileNotFoundError("DB master non trovato.")

    wb = load_workbook(db_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    num_col = len(headers)
    
    # Ensure row_values has the correct length
    row_list = [str(v).strip() for v in row_values]
    if len(row_list) > num_col:
        row_list = row_list[:num_col]
    elif len(row_list) < num_col:
        row_list += [""] * (num_col - len(row_list))

    ws.append(row_list)
    wb.save(db_path)
    return True

# --- GESTIONE TEMPLATE (CONFIG FILE) ---

def carica_templates():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return {
                "convertita": data.get("convertita", DEFAULT_TEMPLATE_CONVERTITA),
                "non_convertita": data.get("non_convertita", DEFAULT_TEMPLATE_NON_CONV),
            }
        else:
            # Crea il file se non esiste con i default
            salva_templates(DEFAULT_TEMPLATE_CONVERTITA, DEFAULT_TEMPLATE_NON_CONV)
            return carica_templates()
    except Exception:
        return {
            "convertita": DEFAULT_TEMPLATE_CONVERTITA,
            "non_convertita": DEFAULT_TEMPLATE_NON_CONV,
        }

def salva_templates(convertita_text, non_conv_text):
    data = {
        "convertita": convertita_text,
        "non_convertita": non_conv_text,
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# --- API ENDPOINTS (FASTAPI) ---

app = FastAPI()

@app.get("/api/data")
async def get_data():
    """Restituisce le colonne e tutte le righe dal DB."""
    try:
        colonne, righe = carica_dati_da_file(DB_PATH)
        return {"colonne": colonne, "righe": righe, "db_name": os.path.basename(DB_PATH)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore caricamento dati: {e}")

@app.post("/api/import")
async def import_file(file: UploadFile = File(...)):
    """Importa righe da un file Excel sorgente."""
    try:
        content = await file.read()
        righe_copiate = importa_dati_da_buffer(DB_PATH, content)
        return {"messaggio": f"Importazione completata. Righe copiate: {righe_copiate}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore importazione: {e}")

@app.post("/api/merge")
async def merge_files(files: list[UploadFile] = File(...)):
    """Deduplica e unisce più file Excel nel DB master."""
    try:
        file_contents = [await f.read() for f in files]
        files_trovati, nuovi = unisci_file_lista(DB_PATH, file_contents)
        return {"messaggio": f"Unione completata. File elaborati: {files_trovati}, Nuove righe aggiunte: {nuovi}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore unione: {e}")

@app.post("/api/row/status")
async def set_row_status(data: dict):
    """Aggiorna lo stato CONVERTITA/NON CONVERTITA di una riga."""
    try:
        riga_excel = data.get('riga_excel')
        convertita = data.get('convertita', False)
        non_convertita = data.get('non_convertita', False)
        pulisci = data.get('pulisci', False)
        
        aggiorna_stato_riga(DB_PATH, riga_excel, convertita, non_convertita, pulisci)
        return {"messaggio": f"Stato riga {riga_excel} aggiornato."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore aggiornamento stato: {e}")

@app.post("/api/row/add")
async def add_manual_row(data: dict):
    """Aggiunge una riga al DB con i valori specificati."""
    try:
        row_values = data.get('values') 
        if not row_values or not isinstance(row_values, list):
            raise ValueError("Valori riga non validi.")
            
        aggiungi_riga_manuale(DB_PATH, row_values)
        return {"messaggio": "Riga aggiunta manualmente con successo."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore aggiunta riga: {e}")

@app.get("/api/templates")
async def get_templates():
    """Restituisce i template di copia."""
    try:
        return carica_templates()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore caricamento template: {e}")

@app.post("/api/templates/save")
async def save_templates_api(data: dict):
    """Salva i template di copia."""
    try:
        conv = data.get('convertita')
        non_conv = data.get('non_convertita')
        if not conv or not non_conv:
            raise ValueError("I testi dei template non possono essere vuoti.")
            
        salva_templates(conv, non_conv)
        return {"messaggio": "Template salvati con successo."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore salvataggio template: {e}")